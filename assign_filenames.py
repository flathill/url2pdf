#!/usr/bin/env python3
"""assign_filenames.py – F列のURLに通し番号ファイル名を割り当て、
   J列に書き込み、G列・H列の [URL|テキスト] を [ファイル名|テキスト] に置換する。
   URLのドメイン/パスからプレフィックスを自動判別。
   マッチしないURLは DEFAULT グループに振り分け。
   出力は常に別ファイル名（_named 付き）に保存。

使い方:
    python assign_filenames.py input.xlsx
    python assign_filenames.py input.xlsx --rules "zabbix=ZBX,veeam=VEM"
    python assign_filenames.py input.xlsx --rules "zabbix=ZBX,veeam=VEM" --default-prefix GEN
    python assign_filenames.py input.xlsx --dry-run
"""

import argparse, re, sys, datetime
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook


# ── デフォルトのURL→プレフィックス判別ルール ──────────
DEFAULT_RULES = "zabbix=ZBX,veeam=VEM"
DEFAULT_PREFIX = "HRV"


def parse_rules(rules_str):
    """'zabbix=ZBX,veeam=VEM' 形式をパース"""
    rules = []
    for pair in rules_str.split(","):
        pair = pair.strip()
        if not pair or "=" not in pair:
            continue
        keyword, prefix = pair.split("=", 1)
        keyword = keyword.strip().lower()
        prefix = prefix.strip()
        if keyword and prefix:
            rules.append((keyword, prefix))
    return rules


def url_to_prefix(url, rules, default_prefix):
    """URLにマッチするルールからプレフィックスを返す"""
    url_lower = url.lower()
    for keyword, prefix in rules:
        if keyword in url_lower:
            return prefix
    return default_prefix


def normalize_url(url):
    """比較用にURLを正規化（末尾スラッシュ、フラグメント除去）"""
    url = url.strip().rstrip("/")
    url = re.sub(r'#.*$', '', url)
    return url


def extract_urls_from_cell(cell_value):
    """F列のセル値からURLを抽出"""
    if not cell_value:
        return []
    urls = re.findall(
        r'https?://[^\s\u3000-\u9fff\uff00-\uffef（）「」【】、。\n\u200b]+',
        str(cell_value)
    )
    return [u.rstrip('.,;)') for u in urls if u]


def replace_urls_in_evidence(cell_value, url_to_filename):
    """G列・H列の [URL|テキスト] を [ファイル名|テキスト] に置換"""
    if not cell_value:
        return cell_value

    text = str(cell_value)

    def replacer(m):
        inner = m.group(1)
        if '|' not in inner:
            return m.group(0)
        left, right = inner.split('|', 1)
        left_stripped = left.strip()
        if left_stripped.startswith('http://') or left_stripped.startswith('https://'):
            norm = normalize_url(left_stripped)
            if norm in url_to_filename:
                return f"[{url_to_filename[norm]}|{right}]"
        return m.group(0)

    return re.sub(r'\[([^]]+)\]', replacer, text)


def generate_output_path(input_path):
    """入力ファイル名から出力ファイル名を生成: xxx.xlsx → xxx_named_タイムスタンプ.xlsx"""
    p = Path(input_path)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return p.parent / f"{p.stem}_named_{timestamp}{p.suffix}"


def main():
    parser = argparse.ArgumentParser(
        description="F列URLに通し番号ファイル名を割り当て、G/H列を置換、J列に書き込み")
    parser.add_argument("excel", type=Path, help="入力Excelファイル")
    parser.add_argument("-o", "--output", type=Path, default=None,
                        help="出力Excelファイル（省略時は自動生成）")
    parser.add_argument("--rules", default=DEFAULT_RULES,
                        help="URL→プレフィックスのルール "
                             f"(デフォルト: '{DEFAULT_RULES}')")
    parser.add_argument("--default-prefix", default=DEFAULT_PREFIX,
                        help=f"どのルールにもマッチしないURLのプレフィックス "
                             f"(デフォルト: {DEFAULT_PREFIX})")
    parser.add_argument("--start", type=int, default=1,
                        help="各プレフィックスの通し番号開始値（デフォルト: 1）")
    parser.add_argument("--digits", type=int, default=3,
                        help="通し番号の桁数（デフォルト: 3）")
    parser.add_argument("--url-col", default="F",
                        help="URL列（デフォルト: F）")
    parser.add_argument("--evidence-cols", default="G,H",
                        help="証跡列（カンマ区切り、デフォルト: G,H）")
    parser.add_argument("--name-col", default="J",
                        help="ファイル名書き込み列（デフォルト: J）")
    parser.add_argument("--dry-run", action="store_true",
                        help="変更せず結果のみ表示")
    args = parser.parse_args()

    output_path = args.output or generate_output_path(args.excel)
    evidence_cols = [c.strip() for c in args.evidence_cols.split(",")]
    rules = parse_rules(args.rules)
    default_prefix = args.default_prefix

    # 列文字→列番号
    def col_to_num(c):
        result = 0
        for ch in c.upper():
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result

    url_col_num = col_to_num(args.url_col)
    evidence_col_nums = [col_to_num(c) for c in evidence_cols]
    name_col_num = col_to_num(args.name_col)

    print("=" * 60)
    print("  assign_filenames – URL→ファイル名割り当て")
    print("=" * 60)
    print(f"  入力: {args.excel}")
    print(f"  出力: {output_path}")
    print(f"  ルール:")
    for kw, px in rules:
        print(f"    URL に '{kw}' を含む → {px}xxx")
    print(f"    マッチなし（DEFAULT） → {default_prefix}xxx")
    print(f"  開始番号: {args.start}  桁数: {args.digits}")
    print("=" * 60)

    # ── Excel読み込み ──
    wb = load_workbook(str(args.excel))
    ws = wb.active

    # ── パス1: 全URLを収集、プレフィックス別に通し番号を割り当て ──
    url_to_filename = {}
    url_order = []
    prefix_counters = defaultdict(lambda: args.start)

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row, url_col_num).value
        urls = extract_urls_from_cell(cell_value)
        for url in urls:
            norm = normalize_url(url)
            if norm not in url_to_filename:
                prefix = url_to_prefix(url, rules, default_prefix)
                num = prefix_counters[prefix]
                fname = f"{prefix}{num:0{args.digits}d}"
                prefix_counters[prefix] = num + 1
                url_to_filename[norm] = fname
                url_order.append((norm, url, fname, prefix))

    # ── 集計・マッピング表示 ──
    prefix_counts = defaultdict(int)
    for _, _, _, px in url_order:
        prefix_counts[px] += 1

    print(f"\n  URL検出: {len(url_order)} 件（ユニーク）")
    for px in sorted(prefix_counts.keys()):
        label = f"{px} (DEFAULT)" if px == default_prefix else f"{px}"
        print(f"    {label}: {prefix_counts[px]} 件")
    print()

    # プレフィックス別にグループ表示
    by_prefix = defaultdict(list)
    for norm, orig_url, fname, prefix in url_order:
        by_prefix[prefix].append((fname, orig_url))

    for px in sorted(by_prefix.keys()):
        items = by_prefix[px]
        label = f"{px}グループ (DEFAULT)" if px == default_prefix else f"{px}グループ"
        print(f"  ┌─ {label} ({len(items)}件)")
        for fname, orig_url in items:
            short_url = (orig_url[:55] + "...") if len(orig_url) > 58 else orig_url
            print(f"  │  {fname}  ←  {short_url}")
        print(f"  └─")
        print()

    if args.dry_run:
        print("  [dry-run] 変更は保存されません。")
        print(f"  （実行時の出力先: {output_path}）")
        wb.close()
        return

    # ── パス2: J列書き込み + G/H列置換 ──
    rows_modified = 0
    replacements_gh = 0

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row, url_col_num).value
        urls = extract_urls_from_cell(cell_value)
        if not urls:
            continue

        # J列
        filenames_for_row = []
        for url in urls:
            norm = normalize_url(url)
            if norm in url_to_filename:
                filenames_for_row.append(url_to_filename[norm])
        if filenames_for_row:
            ws.cell(row, name_col_num).value = "\n".join(filenames_for_row)

        # G列・H列
        for ec in evidence_col_nums:
            old_value = ws.cell(row, ec).value
            if old_value:
                new_value = replace_urls_in_evidence(old_value, url_to_filename)
                if new_value != str(old_value):
                    ws.cell(row, ec).value = new_value
                    replacements_gh += 1

        rows_modified += 1

    # ── 別ファイルに保存 ──
    wb.save(str(output_path))
    wb.close()

    print(f"  処理完了:")
    print(f"    URL（ユニーク）: {len(url_order)} 件")
    for px in sorted(prefix_counts.keys()):
        label = f"{px} (DEFAULT)" if px == default_prefix else f"{px}"
        print(f"      {label}: {prefix_counts[px]} 件")
    print(f"    J列更新: {rows_modified} 行")
    print(f"    G/H列置換: {replacements_gh} 箇所")
    print(f"    保存先: {output_path}")
    print()
    print(f"  ※ 元ファイル '{args.excel.name}' は変更されていません。")
    print("=" * 60)


if __name__ == "__main__":
    main()
