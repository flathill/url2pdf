#!/usr/bin/env python3
"""url2pdf.py - Excel(A-J列)からURLを読み取り、証跡PDFを保存する"""

import argparse, re, sys, os, time, datetime
from pathlib import Path
from multiprocessing import Process, Queue, Manager

# ---------------------------------------------------------------------------
# Excel 解析
# ---------------------------------------------------------------------------
def extract_urls(cell_value):
    """F列のセル値からURLを抽出"""
    if not cell_value:
        return []
    urls = re.findall(
        r'https?://[^\s\u3000-\u9fff\uff00-\uffef（）「」【】、。\n\u200b]+',
        str(cell_value)
    )
    def _smart_rstrip(u):
        while u and u[-1] in '.,;':
            u = u[:-1]
        while u.endswith(')') and u.count(')') > u.count('('):
            u = u[:-1]
        return u
    return [_smart_rstrip(u) for u in urls if u]

def parse_filenames(j_value):
    """J列のセル値をファイル名リストに変換"""
    if not j_value:
        return []
    return [fn.strip() for fn in str(j_value).split('\n') if fn.strip()]

def load_download_tasks(excel_path, url_col_idx=6, name_col_idx=9):
    """Excelから (url, filename) ペアを重複排除して返す"""
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active

    seen_filenames = set()
    tasks = []

    for row in range(2, ws.max_row + 1):
        f_value = ws.cell(row, url_col_idx).value
        j_value = ws.cell(row, name_col_idx).value

        urls = extract_urls(f_value)
        filenames = parse_filenames(j_value)

        for url, fname in zip(urls, filenames):
            if fname in seen_filenames:
                continue
            seen_filenames.add(fname)
            tasks.append({"url": url, "filename": fname + ".pdf"})

    wb.close()
    return tasks

# ---------------------------------------------------------------------------
# PDF 存在チェック
# ---------------------------------------------------------------------------
def check_existing(path, min_size):
    if not path.exists():
        return "new"
    if path.stat().st_size >= min_size:
        return "skip"
    return "redownload"

# ---------------------------------------------------------------------------
# Cookie バナー除去
# ---------------------------------------------------------------------------
def try_dismiss_cookie_banner(page):
    selectors = [
        '#onetrust-accept-btn-handler',
        '#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll',
        '.cookie-accept', '.accept-cookies', '#accept-cookies',
        '[data-testid="cookie-accept"]', '.cc-accept', '.cc-btn.cc-dismiss',
        '#gdpr-cookie-accept', '.js-cookie-consent-agree',
        'button[aria-label="Accept cookies"]',
        'button[aria-label="Accept all cookies"]',
    ]
    for sel in selectors:
        try:
            btn = page.query_selector(sel)
            if btn and btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(1000)
                return True
        except Exception:
            pass

    button_texts = [
        "Accept", "Accept All", "Accept all", "Accept Cookies",
        "Allow All", "Allow all", "I agree", "Agree", "OK", "Got it",
        "Understood", "Continue", "Consent",
        "同意する", "すべて許可", "許可", "承認",
    ]
    for text in button_texts:
        try:
            btn = page.get_by_role("button", name=text, exact=False)
            if btn.count() > 0 and btn.first.is_visible():
                btn.first.click(timeout=2000)
                page.wait_for_timeout(1000)
                return True
        except Exception:
            pass
    return False

def remove_cookie_overlay_by_css(page):
    page.evaluate("""() => {
        const sels = [
            '#onetrust-banner-sdk', '#onetrust-consent-sdk',
            '#CybotCookiebotDialog', '.cookie-banner', '.cookie-consent',
            '.cc-window', '#gdpr-cookie-notice', '.privacy-banner',
            '[class*="cookie"]', '[id*="cookie-banner"]',
            '[class*="consent"]', '[id*="consent"]',
        ];
        sels.forEach(s => {
            document.querySelectorAll(s).forEach(el => {
                el.style.display = 'none';
                el.style.visibility = 'hidden';
            });
        });
        document.querySelectorAll('.modal-backdrop, .overlay').forEach(el => {
            el.style.display = 'none';
        });
        document.body.style.overflow = 'auto';
        document.documentElement.style.overflow = 'auto';
    }""")

# ---------------------------------------------------------------------------
# 画像読み込み + スクロール
# ---------------------------------------------------------------------------
def scroll_and_load_images(page):
    page.evaluate("""() => {
        document.querySelectorAll('img[loading="lazy"]').forEach(img => {
            img.loading = 'eager';
        });
        document.querySelectorAll('img[data-src]').forEach(img => {
            if (!img.src || img.src.includes('placeholder') || img.src.includes('data:')) {
                img.src = img.dataset.src;
            }
        });
    }""")
    try:
        vh = page.evaluate("() => window.innerHeight || 900")
    except Exception:
        vh = 900
    step = int(vh * 0.8)
    try:
        total_h = page.evaluate("() => Math.max(document.body.scrollHeight, document.documentElement.scrollHeight)")
    except Exception:
        total_h = 5000
    pos = 0
    while pos < total_h:
        pos += step
        page.evaluate(f"() => window.scrollTo(0, {pos})")
        page.wait_for_timeout(400)
    page.evaluate("() => window.scrollTo(0, 0)")
    page.wait_for_timeout(500)

# ---------------------------------------------------------------------------
# Docusaurus / SPA 対策: CSS レイアウト修正
# ---------------------------------------------------------------------------
def fix_layout_for_pdf(page):
    page.evaluate("""() => {
        const all = document.querySelectorAll('*');
        for (const el of all) {
            const cs = window.getComputedStyle(el);
            if (cs.overflow === 'hidden' || cs.overflowY === 'hidden') {
                el.style.overflow = 'visible';
                el.style.overflowY = 'visible';
            }
            if (cs.height && cs.height.includes('vh')) {
                el.style.height = 'auto';
            }
            if (cs.maxHeight && cs.maxHeight !== 'none') {
                el.style.maxHeight = 'none';
            }
            if (cs.position === 'sticky') {
                el.style.position = 'relative';
            }
            if (cs.position === 'fixed') {
                el.style.position = 'absolute';
            }
        }
        document.documentElement.style.height = 'auto';
        document.documentElement.style.overflow = 'visible';
        document.body.style.height = 'auto';
        document.body.style.overflow = 'visible';

        // Docusaurus specific
        const ds = document.querySelector('#__docusaurus');
        if (ds) {
            ds.style.height = 'auto';
            ds.style.overflow = 'visible';
            ds.style.minHeight = 'auto';
        }
        document.querySelectorAll('[class*="docPage"], [class*="docMainContainer"], [class*="docRoot"], main, article').forEach(el => {
            el.style.height = 'auto';
            el.style.overflow = 'visible';
            el.style.maxHeight = 'none';
        });
    }""")

def prepare_for_pdf(page):
    page.evaluate("() => window.scrollTo(0, 0)")
    page.wait_for_timeout(500)
    page.evaluate("""() => {
        const imgs = Array.from(document.images).filter(i => !i.complete);
        return Promise.all(imgs.map(i => new Promise(r => {
            i.onload = i.onerror = r;
            setTimeout(r, 5000);
        })));
    }""")
    page.wait_for_timeout(500)

# ---------------------------------------------------------------------------
# ヘッダー / フッター
# ---------------------------------------------------------------------------
def build_header_footer_style():
    return ("font-family: Arial, sans-serif; font-size: 9px; width: 100%; "
            "padding: 0 10mm; box-sizing: border-box; color: #444;")

# ---------------------------------------------------------------------------
# Worker プロセス
# ---------------------------------------------------------------------------
def worker_process(task_queue, result_queue, output_dir, wait_seconds, paper_format, landscape, retry_count):
    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(
            viewport={"width": 1280, "height": 900},
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
        )
        page = context.new_page()

        while True:
            task = task_queue.get()
            if task is None:
                break

            url = task["url"]
            filename = task["filename"]
            output_path = Path(output_dir) / filename
            status = "ok"
            detail = ""

            for attempt in range(1 + retry_count):
                try:
                    try:
                        page.goto(url, wait_until="networkidle", timeout=60000)
                    except Exception:
                        page.goto(url, wait_until="domcontentloaded", timeout=60000)

                    page.wait_for_timeout(1500)

                    # Cookie
                    try_dismiss_cookie_banner(page)
                    remove_cookie_overlay_by_css(page)

                    # Images + scroll
                    scroll_and_load_images(page)

                    # Layout fix
                    fix_layout_for_pdf(page)

                    # Extra wait
                    page.wait_for_timeout(wait_seconds * 1000)

                    # Final cleanup
                    remove_cookie_overlay_by_css(page)
                    prepare_for_pdf(page)

                    # Header / Footer
                    title = page.title() or "(タイトルなし)"
                    accessed_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S JST")
                    safe_title = title.replace("'", "&#39;").replace('"', "&quot;")[:80]
                    safe_url = url.replace("'", "&#39;").replace('"', "&quot;")

                    header_html = (
                        f"<div style='{build_header_footer_style()} border-bottom:1px solid #ccc; "
                        f"padding-bottom:2mm; display:flex; justify-content:space-between;'>"
                        f"<span style='flex:1;text-align:left;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;'>{safe_title}</span>"
                        f"<span style='flex:1;text-align:right;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;'>{safe_url}</span>"
                        f"</div>"
                    )
                    footer_html = (
                        f"<div style='{build_header_footer_style()} border-top:1px solid #ccc; "
                        f"padding-top:2mm; display:flex; justify-content:space-between;'>"
                        f"<span>取得日時: {accessed_at}  |  {filename}</span>"
                        f"<span><span class='pageNumber'></span> / <span class='totalPages'></span> ページ</span>"
                        f"</div>"
                    )

                    page.pdf(
                        path=str(output_path),
                        format=paper_format,
                        landscape=landscape,
                        print_background=True,
                        display_header_footer=True,
                        header_template=header_html,
                        footer_template=footer_html,
                        margin={"top": "25mm", "bottom": "20mm", "left": "10mm", "right": "10mm"},
                    )
                    status = "ok"
                    detail = f"{output_path.stat().st_size / 1024:.0f} KB"
                    break

                except Exception as e:
                    status = "error"
                    detail = str(e)[:120]
                    if attempt < retry_count:
                        time.sleep(3)

            result_queue.put({"filename": filename, "url": url, "status": status, "detail": detail})

        browser.close()

# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Excel(A-J列)からURLを読み取り証跡PDFを保存")
    parser.add_argument("excel", type=Path, help="入力Excelファイル")
    parser.add_argument("-o", "--output-dir", type=Path, default=Path("."), help="出力ディレクトリ")
    parser.add_argument("-w", "--wait", type=int, default=5, help="追加待機秒数")
    parser.add_argument("-f", "--format", default="A4", help="用紙サイズ")
    parser.add_argument("-l", "--landscape", action="store_true", help="横向き")
    parser.add_argument("-r", "--retry", type=int, default=1, help="リトライ回数")
    parser.add_argument("-j", "--workers", type=int, default=5, help="並列ワーカー数")
    parser.add_argument("--min-size", type=int, default=60000, help="正常PDF最小バイト数")
    parser.add_argument("--force", action="store_true", help="既存PDF強制上書き")
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("  url2pdf - 証跡PDF生成ツール (新Excel構造対応)")
    print("=" * 60)
    print(f"  Excel       : {args.excel}")
    print(f"  出力先      : {args.output_dir}")
    print(f"  ワーカー数  : {args.workers}")

    # タスク読み込み
    all_tasks = load_download_tasks(args.excel)
    print(f"  URL総数     : {len(all_tasks)}")

    # スキップ判定
    download_targets = []
    skip_count = 0
    for t in all_tasks:
        out = args.output_dir / t["filename"]
        if args.force:
            download_targets.append(t)
        else:
            st = check_existing(out, args.min_size)
            if st == "skip":
                skip_count += 1
            else:
                download_targets.append(t)

    print(f"  スキップ    : {skip_count}")
    print(f"  取得対象    : {len(download_targets)}")
    print("=" * 60)

    if not download_targets:
        print("  取得対象なし。終了します。")
        return

    start_time = time.time()

    task_queue = Queue()
    result_queue = Queue()

    for t in download_targets:
        task_queue.put(t)

    num_workers = min(args.workers, len(download_targets))
    for _ in range(num_workers):
        task_queue.put(None)  # sentinel

    workers = []
    for _ in range(num_workers):
        p = Process(target=worker_process, args=(
            task_queue, result_queue, str(args.output_dir),
            args.wait, args.format, args.landscape, args.retry
        ))
        p.start()
        workers.append(p)

    # 結果収集
    results = []
    for i in range(len(download_targets)):
        r = result_queue.get()
        mark = "OK" if r["status"] == "ok" else "NG"
        print(f"  [{i+1}/{len(download_targets)}] {r['filename']} ... {mark}  {r['detail']}")
        results.append(r)

    for p in workers:
        p.join()

    elapsed = time.time() - start_time
    ok_count = sum(1 for r in results if r["status"] == "ok")
    ng_count = sum(1 for r in results if r["status"] != "ok")

    print("=" * 60)
    print(f"  完了: 成功 {ok_count} 件 / 失敗 {ng_count} 件 / スキップ {skip_count} 件")
    print(f"  所要時間: {elapsed:.1f} 秒")
    print(f"  出力先: {args.output_dir}")
    print("=" * 60)

    if ng_count > 0:
        print("\n  ■ 失敗一覧:")
        for r in results:
            if r["status"] != "ok":
                print(f"    {r['filename']} : {r['detail']}")

if __name__ == "__main__":
    main()
