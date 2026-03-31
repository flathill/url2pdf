#!/usr/bin/env python3
"""pdf_annotate_v13.9 – 証跡PDFアノテーション
   v13.6ベース + pymupdf 1.25+ 対応（border_color を xref_set_key で設定）
   全てFreeTextアノテーション（編集・移動可能）
   + clean_contents
   + 項番の重複排除（同一ページ内で同じ項番は1回だけ）
   + 日本語翻訳は常に出力（同一項番でも各ハイライトごとに）
   + 処理完了後にExcel行（要件）単位のサマリレポート表示
   + 重複排除でスキップされた行にも結果を反映
   + コメント付きページ±1ページを抽出（不要ページ削除方式でサイズ抑制）
   + 出力先に既存ファイルがあればスキップ（再処理防止）
"""

import argparse, os, re, shutil, sys, time, json
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
import pymupdf


# ── ユーティリティ ──────────────────────────────────

def norm(s):
    return re.sub(r'\s+', ' ', s).strip().lower()


def parse_evidence(cell_value):
    if not cell_value:
        return []
    results = []
    for m in re.finditer(r'\[([^]]+)\]', str(cell_value)):
        inner = m.group(1)
        if '|' in inner:
            fname, text = inner.split('|', 1)
            results.append((fname.strip(), text.strip()))
    return results


def pdf_fingerprint(pdf_path):
    st = os.stat(pdf_path)
    return f"{st.st_mtime:.6f}_{st.st_size}"


def set_annot_border_color(doc, annot, color_tuple, width=0.5):
    """pymupdf 1.25+ 対応: update()後にAPストリーム内の枠線描画を書き換え"""
    try:
        xref = annot.xref
        # アノテーション属性に色と枠線を設定
        c_str = "[" + " ".join(f"{c:.4f}" for c in color_tuple) + "]"
        doc.xref_set_key(xref, "C", c_str)
        bs_str = f"<</W {width} /S /S>>"
        doc.xref_set_key(xref, "BS", bs_str)
        # APストリーム内に枠線描画を追加
        ap_str = doc.xref_get_key(xref, "AP")
        if ap_str[0] == "dict":
            # /AP << /N xref >> の形式からNormalのxrefを取得
            import re as _re
            m = _re.search(r"(\d+) 0 R", ap_str[1])
            if m:
                ap_n_xref = int(m.group(1))
                stream = doc.xref_stream(ap_n_xref)
                if stream:
                    # 矩形を取得してストロークコマンドを生成
                    r = annot.rect
                    w2 = width / 2
                    stroke_cmd = (
                        f"q {color_tuple[0]:.4f} {color_tuple[1]:.4f} {color_tuple[2]:.4f} RG "
                        f"{width} w "
                        f"{w2:.2f} {w2:.2f} {r.width - width:.2f} {r.height - width:.2f} re S Q\n"
                    )
                    if isinstance(stream, bytes):
                        new_stream = stream + stroke_cmd.encode()
                    else:
                        new_stream = stream.encode() + stroke_cmd.encode()
                    doc.xref_set_stream(ap_n_xref, new_stream)
    except Exception:
        pass


# ── テキストインデックス ─────────────────────────────

def build_text_index(pdf_dir, cache_file=None):
    pdf_files = sorted(Path(pdf_dir).glob("*.pdf"))
    total = len(pdf_files)
    cached = {}
    if cache_file and os.path.exists(cache_file):
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                cached = json.load(f)
            print(f"  キャッシュ: {len(cached.get('data', {}))} ファイル")
        except Exception:
            cached = {}
    cached_data = cached.get("data", {})
    cached_fp = cached.get("fingerprints", {})
    idx = {}
    reindexed = from_cache = 0
    print(f"  インデックス構築中 ({total} ファイル)...", flush=True)
    t0 = time.time()
    for i, p in enumerate(pdf_files):
        stem = p.stem
        fp = pdf_fingerprint(str(p))
        if stem in cached_data and cached_fp.get(stem) == fp:
            idx[stem] = [(pg[0], pg[1]) for pg in cached_data[stem]]
            from_cache += 1
        else:
            pages = []
            try:
                doc = pymupdf.open(str(p))
                for pi in range(len(doc)):
                    txt = norm(doc[pi].get_text("text"))
                    pages.append((pi, txt))
                doc.close()
            except Exception:
                pass
            idx[stem] = pages
            reindexed += 1
        if (i + 1) % 20 == 0 or i == total - 1:
            print(f"    ... {i+1}/{total} (cache:{from_cache} new:{reindexed})", flush=True)
    print(f"  完了 ({time.time()-t0:.1f}秒)")
    if cache_file:
        new_cache = {"fingerprints": {}, "data": {}}
        for p in pdf_files:
            s = p.stem
            new_cache["fingerprints"][s] = pdf_fingerprint(str(p))
            new_cache["data"][s] = idx.get(s, [])
        try:
            with open(cache_file, "w", encoding="utf-8") as f:
                json.dump(new_cache, f, ensure_ascii=False)
        except Exception:
            pass
    return idx


def search_in_index(index_pages, search_text):
    ns = norm(search_text)
    for pi, pt in index_pages:
        if ns in pt:
            return pi
    for ratio in [0.8, 0.6, 0.4]:
        sub = ns[:max(15, int(len(ns) * ratio))]
        for pi, pt in index_pages:
            if sub in pt:
                return pi
    for clen in [30, 15]:
        sub = ns[:clen]
        if len(sub) >= 10:
            for pi, pt in index_pages:
                if sub in pt:
                    return pi
    return -1


def fallback_search(text_index, search_text, exclude_stem):
    for stem, pages in text_index.items():
        if stem == exclude_stem:
            continue
        pi = search_in_index(pages, search_text)
        if pi >= 0:
            return (stem, pi)
    return None


# ── 定数 ────────────────────────────────────────────

LABEL_FS = 7
JP_FS = 6


# ── 矩形ユーティリティ ──────────────────────────────

def rects_overlap(r1, r2):
    return not (r1.x1 <= r2.x0 or r2.x1 <= r1.x0 or
                r1.y1 <= r2.y0 or r2.y1 <= r1.y0)


def nudge_rect(rect, occupied, page_height):
    candidate = pymupdf.Rect(rect)
    shift = max(candidate.height + 1, 5)
    for _ in range(30):
        collision = any(rects_overlap(candidate, o) for o in occupied)
        if not collision:
            return candidate
        candidate.y0 += shift
        candidate.y1 += shift
        if candidate.y0 > page_height - 10:
            return pymupdf.Rect(rect)
    return pymupdf.Rect(rect)


# ── PDF処理 ──────────────────────────────────────────

def process_pdf(doc, page_annotations):
    """ページごとにアノテーションを追加。"""
    results = []

    for pi, annots in sorted(page_annotations.items()):
        page = doc[pi]
        page.clean_contents(sanitize=False)
        pw = page.rect.width
        ph = page.rect.height
        occupied = []
        drawn_labels_on_page = set()

        for search_text, item_no, jp_text, row_idx in annots:
            quads = page.search_for(search_text, quads=True)
            if not quads:
                for clen in [80, 50, 30, 20]:
                    sub = search_text[:clen] if len(search_text) > clen else search_text
                    quads = page.search_for(sub, quads=True)
                    if quads:
                        break

            if not quads:
                results.append((False, row_idx, item_no, pi))
                continue

            ann = page.add_highlight_annot(quads=quads)
            ann.set_colors(stroke=(1, 1, 0))
            ann.update()

            first_q = quads[0]
            last_q = quads[-1]
            first_r = first_q.rect if hasattr(first_q, "rect") else pymupdf.Rect(first_q)
            last_r = last_q.rect if hasattr(last_q, 'rect') else pymupdf.Rect(last_q)
            hl_y0 = first_r.y0
            hl_x0 = first_r.x0
            hl_y1 = last_r.y1

            if item_no not in drawn_labels_on_page:
                label_h = LABEL_FS * 2.5
                label_w = max(len(item_no) * 4.5 + 8, 50)
                label_y0 = hl_y0 - label_h
                if label_y0 < 0:
                    label_y0 = 0
                lx1 = hl_x0 - 2
                lx0 = lx1 - label_w
                if lx0 < 0:
                    lx0 = 0
                    lx1 = lx0 + label_w

                label_rect = pymupdf.Rect(lx0, label_y0, lx1, label_y0 + label_h)
                label_rect = nudge_rect(label_rect, occupied, ph)

                a = page.add_freetext_annot(
                    label_rect,
                    item_no,
                    fontsize=LABEL_FS,
                    fontname="japan",
                    text_color=(1, 0, 0),
                    fill_color=(1, 1, 1),
                    align=0,
                )
                a.update()
                set_annot_border_color(doc, a, (1, 0, 0), width=0.5)
                occupied.append(pymupdf.Rect(label_rect))
                drawn_labels_on_page.add(item_no)

            if jp_text:
                avail_w = pw - hl_x0 - 10
                if avail_w < 80:
                    avail_w = pw - 20
                    jp_x0 = 10
                else:
                    jp_x0 = hl_x0
                chars_per_line = max(1, int(avail_w / (JP_FS * 0.55)))
                n_lines = max(1, -(-len(jp_text) // chars_per_line))
                jp_h = n_lines * (JP_FS * 1.5) + JP_FS
                jp_h = max(jp_h, JP_FS * 2.5)
                if hl_y1 + 1 + jp_h > ph:
                    jp_h = max(ph - hl_y1 - 2, JP_FS * 2.5)

                jp_rect = pymupdf.Rect(
                    jp_x0, hl_y1 + 1,
                    jp_x0 + avail_w, hl_y1 + 1 + jp_h
                )
                jp_rect = nudge_rect(jp_rect, occupied, ph)

                a2 = page.add_freetext_annot(
                    jp_rect,
                    jp_text,
                    fontsize=JP_FS,
                    fontname="japan",
                    text_color=(0, 0, 0.7),
                    fill_color=(1, 1, 1),
                    align=0,
                )
                a2.update()
                set_annot_border_color(doc, a2, (0.7, 0.7, 1), width=0.5)
                occupied.append(pymupdf.Rect(jp_rect))

            results.append((True, row_idx, item_no, pi))

    return results


# ── ページ抽出（不要ページ削除方式）─────────────────

def extract_annotated_pages(output_dir, annotated_pages_map):
    extract_count = 0
    skip_count = 0

    print()
    print("=" * 60)
    print("  ページ抽出（コメント箇所 ± 1ページ）")
    print("=" * 60)

    for stem in sorted(annotated_pages_map.keys()):
        commented_pages = annotated_pages_map[stem]
        if not commented_pages:
            continue

        src_path = os.path.join(output_dir, f"{stem}.pdf")
        extract_path = os.path.join(output_dir, f"{stem}_抽出.pdf")
        tmp_path = os.path.join(output_dir, f"{stem}_抽出.tmp.pdf")

        if os.path.exists(extract_path):
            print(f"    {stem}_抽出.pdf … スキップ（既存）")
            skip_count += 1
            continue

        if not os.path.exists(src_path):
            print(f"    {stem}.pdf … ソースなし")
            continue

        doc = pymupdf.open(src_path)
        total_pages = len(doc)

        keep_set = set()
        for p in commented_pages:
            for offset in (-1, 0, 1):
                candidate = p + offset
                if 0 <= candidate < total_pages:
                    keep_set.add(candidate)

        delete_pages = sorted(
            [p for p in range(total_pages) if p not in keep_set],
            reverse=True
        )

        for p in delete_pages:
            doc.delete_page(p)

        doc.save(tmp_path, garbage=4, deflate=True)
        doc.close()
        shutil.move(tmp_path, extract_path)

        keep_list = sorted(keep_set)
        commented_desc = ", ".join(str(p + 1) for p in sorted(commented_pages))
        extracted_desc = ", ".join(str(p + 1) for p in keep_list)
        print(f"    {stem}_抽出.pdf … {len(keep_list)}/{total_pages}ページ残し "
              f"(p.{extracted_desc}) ← コメント: p.{commented_desc}")
        extract_count += 1

    print(f"  ── 抽出完了: {extract_count}件生成, {skip_count}件スキップ")
    print("=" * 60)

    return extract_count, skip_count

# ── サマリレポート ───────────────────────────────────

def print_summary_report(row_results):
    print()
    print("=" * 76)
    print("  サマリレポート（Excel要件単位）")
    print("=" * 76)
    print(f"  {'行':>4}  {'項番':<14}  {'結果':^4}  マーキング箇所")
    print("-" * 76)

    total_ok = total_fb = total_fail = 0

    for row_idx in sorted(row_results.keys()):
        info = row_results[row_idx]
        item_no = info["item_no"]
        marks = info["marks"]

        ok_marks = [(s, p) for s, p, st in marks if st == "ok"]
        fb_marks = [(s, p) for s, p, st in marks if st == "fb"]
        fail_marks = [(s, p) for s, p, st in marks if st == "fail"]
        dedup_marks = [(s, p) for s, p, st in marks if st == "dedup"]

        total_ok += len(ok_marks)
        total_fb += len(fb_marks)
        total_fail += len(fail_marks)

        parts = []
        for s, p in ok_marks:
            parts.append(f"{s}.pdf({p+1}p)")
        for s, p in fb_marks:
            parts.append(f"{s}.pdf({p+1}p)⚠FB")
        for s, p in dedup_marks:
            parts.append(f"{s}.pdf({p+1}p)※重複")
        for s, p in fail_marks:
            parts.append(f"{s}.pdf ✗未検出")

        has_success = ok_marks or fb_marks or dedup_marks
        if fail_marks and not has_success:
            status = "✗"
        elif fail_marks:
            status = "△"
        elif fb_marks:
            status = "⚠"
        else:
            status = "✓"

        locations = ", ".join(parts) if parts else "（証跡なし）"
        print(f"  {row_idx:>4}  {item_no:<14}  {status:^4}  {locations}")

    print("-" * 76)
    print(f"  合計: ✓成功 {total_ok}件  ⚠FB {total_fb}件  ✗未検出 {total_fail}件")
    print("=" * 76)
    print()
    print("  凡例: ✓=全て成功  ⚠=フォールバック先で発見  △=一部未検出  ✗=全て未検出")
    print("        ※重複=同一テキストが別行で既にマーキング済み（ハイライトは共有）")
    print()


# ── メイン ───────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel")
    parser.add_argument("--input-dir", default="./pdf")
    parser.add_argument("--output-dir", default="./pdf_annotated")
    parser.add_argument("--cache", default=None)
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    cache_file = args.cache or os.path.join(args.input_dir, ".text_index_cache.json")
    text_index = build_text_index(args.input_dir, cache_file=cache_file)

    # ── Excel読み込み ──
    wb = load_workbook(args.excel, read_only=True, data_only=True)
    ws = wb.active
    tasks = {}
    seen_keys = {}
    row_meta = {}
    dedup_links = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 10:
            continue
        item_no = str(row[2] or "").strip()
        if not item_no:
            continue
        row_meta[row_idx] = item_no
        en_entries = parse_evidence(row[6])
        jp_entries = parse_evidence(row[7])
        for i, (fname, en_text) in enumerate(en_entries):
            stem = fname.strip()
            dedup_key = (stem, norm(en_text))
            if dedup_key in seen_keys:
                orig_row = seen_keys[dedup_key]
                dedup_links.append((row_idx, orig_row, stem, en_text))
                continue
            seen_keys[dedup_key] = row_idx
            jp_text = jp_entries[i][1] if i < len(jp_entries) else ""
            tasks.setdefault(stem, []).append((row_idx, item_no, en_text, jp_text))
    wb.close()

    all_stems = set(text_index.keys())
    task_stems = set(tasks.keys())
    total_annot = sum(len(v) for v in tasks.values())

    stems_need_annotate = set()
    stems_already_done = set()
    for stem in task_stems:
        out_path = os.path.join(args.output_dir, f"{stem}.pdf")
        if os.path.exists(out_path):
            stems_already_done.add(stem)
        else:
            stems_need_annotate.add(stem)

    print(f"""============================================================
  pdf_annotate_v13.9 - FreeText方式（全て編集可能）
  項番重複排除あり / 日本語訳は常に出力
  + コメント箇所±1ページ抽出（不要ページ削除方式）
  + pymupdf 1.25+ 対応（border_color を xref_set_key で設定）
============================================================
  Excel: {os.path.basename(args.excel)}
  対象: {len(task_stems)} ファイル / {total_annot} 件
  新規処理: {len(stems_need_annotate)} ファイル
  スキップ（既存）: {len(stems_already_done)} ファイル
============================================================""")

    row_results = {}
    for ri, ino in row_meta.items():
        row_results[ri] = {"item_no": ino, "marks": []}

    ok_count = fail_count = fallback_count = done_count = 0
    skip_annot_count = 0
    fb_tasks = {}
    original_row_results = defaultdict(list)
    annotated_pages_map = defaultdict(set)

    for stem in sorted(task_stems):
        annots = tasks[stem]
        pdf_path = os.path.join(args.input_dir, f"{stem}.pdf")
        out_path = os.path.join(args.output_dir, f"{stem}.pdf")

        if stem in stems_already_done:
            skip_annot_count += 1
            try:
                existing_doc = pymupdf.open(out_path)
                for pi in range(len(existing_doc)):
                    page = existing_doc[pi]
                    if page.annots():
                        annotated_pages_map[stem].add(pi)
                existing_doc.close()
            except Exception:
                pass
            for row_idx, item_no, en_text, jp_text in annots:
                pages_data = text_index.get(stem, [])
                pi = search_in_index(pages_data, en_text)
                if pi >= 0:
                    if row_idx in row_results:
                        row_results[row_idx]["marks"].append((stem, pi, "ok"))
                    original_row_results[row_idx].append((stem, pi, "ok"))
                    ok_count += 1
                else:
                    fb = fallback_search(text_index, en_text, stem)
                    if fb:
                        fb_stem, fb_pi = fb
                        if row_idx in row_results:
                            row_results[row_idx]["marks"].append((fb_stem, fb_pi, "fb"))
                        original_row_results[row_idx].append((fb_stem, fb_pi, "fb"))
                        fallback_count += 1
                    else:
                        if row_idx in row_results:
                            row_results[row_idx]["marks"].append((stem, -1, "fail"))
                        original_row_results[row_idx].append((stem, -1, "fail"))
                        fail_count += 1
            print(f"  [skip] {stem}.pdf … 既存のためスキップ")
            continue

        done_count += 1
        print(f"  [{done_count}/{len(stems_need_annotate)}] {stem}.pdf ({len(annots)})...",
              end="", flush=True)
        t0 = time.time()

        if not os.path.exists(pdf_path):
            print(f" 不存在")
            for r, ino, et, jt in annots:
                if r in row_results:
                    row_results[r]["marks"].append((stem, -1, "fail"))
                original_row_results[r].append((stem, -1, "fail"))
                fail_count += 1
            continue

        doc = pymupdf.open(pdf_path)
        page_annots = {}
        fb_items = []
        for row_idx, item_no, en_text, jp_text in annots:
            pages_data = text_index.get(stem, [])
            pi = search_in_index(pages_data, en_text)
            if pi >= 0:
                page_annots.setdefault(pi, []).append(
                    (en_text, item_no, jp_text, row_idx))
            else:
                fb_items.append((row_idx, item_no, en_text, jp_text))

        results = process_pdf(doc, page_annots)
        for found, row_idx, item_no, pi in results:
            if found:
                if row_idx in row_results:
                    row_results[row_idx]["marks"].append((stem, pi, "ok"))
                original_row_results[row_idx].append((stem, pi, "ok"))
                annotated_pages_map[stem].add(pi)
                ok_count += 1
            else:
                if row_idx in row_results:
                    row_results[row_idx]["marks"].append((stem, pi, "fail"))
                original_row_results[row_idx].append((stem, pi, "fail"))
                fail_count += 1

        doc.save(out_path)
        doc.close()

        for row_idx, item_no, en_text, jp_text in fb_items:
            fb = fallback_search(text_index, en_text, stem)
            if fb:
                fb_stem, fb_pi = fb
                fb_tasks.setdefault(fb_stem, {}).setdefault(fb_pi, []).append(
                    (en_text, item_no, jp_text, row_idx))
                fallback_count += 1
            else:
                if row_idx in row_results:
                    row_results[row_idx]["marks"].append((stem, -1, "fail"))
                original_row_results[row_idx].append((stem, -1, "fail"))
                fail_count += 1

        print(f" ({time.time()-t0:.1f}s)")

    # ── フォールバック先PDFへ追記 ──
    for fb_stem, fb_page_annots in fb_tasks.items():
        fb_path = os.path.join(args.input_dir, f"{fb_stem}.pdf")
        fb_out = os.path.join(args.output_dir, f"{fb_stem}.pdf")
        if os.path.exists(fb_out) and fb_stem in stems_already_done:
            for fb_pi, entries in fb_page_annots.items():
                for en_text, item_no, jp_text, row_idx in entries:
                    if row_idx in row_results:
                        row_results[row_idx]["marks"].append((fb_stem, fb_pi, "fb"))
                    original_row_results[row_idx].append((fb_stem, fb_pi, "fb"))
                    annotated_pages_map[fb_stem].add(fb_pi)
                    ok_count += 1
            continue

        fb_doc = pymupdf.open(fb_out if os.path.exists(fb_out) else fb_path)
        results = process_pdf(fb_doc, fb_page_annots)
        for found, row_idx, item_no, pi in results:
            if found:
                if row_idx in row_results:
                    row_results[row_idx]["marks"].append((fb_stem, pi, "fb"))
                original_row_results[row_idx].append((fb_stem, pi, "fb"))
                annotated_pages_map[fb_stem].add(pi)
                ok_count += 1
        fb_tmp = fb_out + ".tmp"
        fb_doc.save(fb_tmp)
        fb_doc.close()
        shutil.move(fb_tmp, fb_out)

    # ── 重複排除でスキップされた行に結果をコピー ──
    dedup_done = set()
    for skipped_row, orig_row, stem, en_text in dedup_links:
        dedup_id = (skipped_row, stem, norm(en_text))
        if dedup_id in dedup_done:
            continue
        dedup_done.add(dedup_id)
        if skipped_row in row_results:
            for s, p, st in original_row_results.get(orig_row, []):
                if s == stem and (s, p, "dedup") not in row_results[skipped_row]["marks"]:
                    row_results[skipped_row]["marks"].append((s, p, "dedup"))
                    break

    # ── 未参照PDF削除（pdf_annotatedから） ──
    referenced_stems = task_stems | set(fb_tasks.keys())
    del_count = 0
    for pdf_file in sorted(Path(args.output_dir).glob("*.pdf")):
        stem = pdf_file.stem
        if stem.endswith("_抽出"):
            base_stem = stem[:-3]
            if base_stem not in referenced_stems:
                pdf_file.unlink()
                del_count += 1
                print(f"    削除: {pdf_file.name}")
        elif stem not in referenced_stems:
            pdf_file.unlink()
            del_count += 1
            print(f"    削除: {pdf_file.name}")
    copy_count = 0

    # ── サマリー ──
    print(f"""
============================================================
  完了! ✓{ok_count} ⚠FB:{fallback_count} ✗{fail_count}
  PDF: {done_count}新規処理 + {skip_annot_count}スキップ + {del_count}件削除
  出力先: {args.output_dir}
============================================================""")

    print_summary_report(row_results)

    # ── 索引Excel出力（J列） ──
    print()
    print("── 索引Excel出力 ──")
    try:
        from openpyxl import load_workbook as _lwb
        wb2 = _lwb(args.excel)
        ws2 = wb2.active
        idx_count = 0
        for row_idx, info in row_results.items():
            marks = info.get("marks", [])
            index_entries = []
            for s_stem, p, st in marks:
                if p is not None and p >= 0 and st in ("ok", "fb", "dedup"):
                    entry = f"{s_stem}-P{p+1:03d}"
                    if entry not in index_entries:
                        index_entries.append(entry)
            if index_entries:
                ws2.cell(row=row_idx, column=10, value=", ".join(index_entries))
                idx_count += 1
        out_stem = Path(args.excel).stem
        idx_excel = os.path.join(args.output_dir, f"{out_stem}_indexed.xlsx")
        wb2.save(idx_excel)
        print(f"  {idx_count}行にJ列索引を書き込み -> {idx_excel}")
    except Exception as e:
        print(f"  索引Excel出力エラー: {e}")


    extract_count, extract_skip = extract_annotated_pages(
        args.output_dir, dict(annotated_pages_map))

    print(f"""
============================================================
  全処理完了
  アノテーション: {done_count}新規 + {skip_annot_count}既存スキップ
  ページ抽出: {extract_count}件生成 + {extract_skip}件既存スキップ
  出力先: {args.output_dir}
============================================================""")


if __name__ == "__main__":
    main()
