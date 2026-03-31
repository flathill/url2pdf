#!/usr/bin/env python3
"""restore_urls.py - assign_filenames.py の逆操作
   I列のファイル名とF列のURLの対応を構築し、
   G列・H列の [ファイル名|テキスト] を [URL|テキスト] に復元する。
"""

import argparse, re, sys, datetime
from pathlib import Path
from collections import defaultdict

def col_letter_to_idx(letter):
    """A=1, B=2, ..., Z=26"""
    letter = letter.upper().strip()
    result = 0
    for c in letter:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result

def extract_urls(cell_value):
    """セル値からURLを抽出"""
    if not cell_value:
        return []
    urls = re.findall(
        r'https?://[^\s\u3000-\u9fff\uff00-\uffef（）「」【】、。\n\u200b]+',
        str(cell_value)
    )
    return [u.rstrip('.,;)') for u in urls if u]

def parse_filenames(cell_value):
    """I列のセル値をファイル名リストに変換"""
    if not cell_value:
        return []
    return [fn.strip() for fn in str(cell_value).split('\n') if fn.strip()]

def main():
    parser = argparse.ArgumentParser(
        description="assign_filenames.py の逆操作: ファイル名をURLに復元")
    parser.add_argument("excel", type=Path, help="入力Excelファイル（assign_filenames処理済み）")
    parser.add_argument("-o", "--output", type=Path, default=None, help="出力Excelファイル")
    parser.add_argument("--url-col", default="F", help="URL列 (default: F)")
    parser.add_argument("--evidence-cols", default="G,H", help="証跡テキスト列 (default: G,H)")
    parser.add_argument("--name-col", default="I", help="ファイル名列 (default: I)")
    parser.add_argument("--dry-run", action="store_true", help="変更せず結果のみ表示")
    args = parser.parse_args()

    import openpyxl

    url_col = col_letter_to_idx(args.url_col)
    ev_cols = [col_letter_to_idx(c.strip()) for c in args.evidence_cols.split(",")]
    name_col = col_letter_to_idx(args.name_col)

    # ── パス1: ファイル名→URL対応表を構築 ──
    wb = openpyxl.load_workbook(str(args.excel))
    ws = wb.active

    fname_to_url = {}
    rows_scanned = 0

    for row in range(2, ws.max_row + 1):
        f_value = ws.cell(row, url_col).value
        i_value = ws.cell(row, name_col).value

        urls = extract_urls(f_value)
        filenames = parse_filenames(i_value)

        for url, fname in zip(urls, filenames):
            if fname and url:
                if fname not in fname_to_url:
                    fname_to_url[fname] = url
        rows_scanned += 1

    # ── 表示 ──
    groups = defaultdict(list)
    for fname, url in sorted(fname_to_url.items()):
        prefix = re.match(r'^([A-Z]+)', fname)
        grp = prefix.group(1) if prefix else "OTHER"
        groups[grp].append((fname, url))

    print("=" * 60)
    print("  restore_urls – ファイル名→URL復元")
    print("=" * 60)
    print(f"  入力: {args.excel}")
    print(f"  スキャン: {rows_scanned} 行")
    print(f"  ファイル名→URL対応: {len(fname_to_url)} 件")
    print()

    for grp in sorted(groups.keys()):
        items = groups[grp]
        print(f"  ┌─ {grp}グループ ({len(items)}件)")
        for fname, url in items:
            print(f"  │  {fname}  →  {url}")
        print(f"  └─")
        print()

    if args.dry_run:
        print("  ※ dry-run モード: Excelは変更されません。")
        print("=" * 60)
        wb.close()
        return

    # ── パス2: G列・H列の [ファイル名|テキスト] を [URL|テキスト] に置換 ──
    replace_count = 0
    clear_count = 0

    for row in range(2, ws.max_row + 1):
        # G列・H列の置換
        for ec in ev_cols:
            cell = ws.cell(row, ec)
            val = cell.value
            if not val:
                continue
            original = str(val)
            new_val = original

            # [ファイル名|テキスト] → [URL|テキスト]
            def replace_fname(m):
                fname = m.group(1).strip()
                text = m.group(2)
                if fname in fname_to_url:
                    return f"[{fname_to_url[fname]}|{text}]"
                return m.group(0)

            new_val = re.sub(r'\[([A-Z]+\d+)\|([^\]]*)\]', replace_fname, new_val)

            if new_val != original:
                cell.value = new_val
                replace_count += 1

        # I列をクリア
        i_cell = ws.cell(row, name_col)
        if i_cell.value:
            i_cell.value = None
            clear_count += 1

    # ── 保存 ──
    if args.output:
        out_path = args.output
    else:
        stem = args.excel.stem
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = args.excel.parent / f"{stem}_restored_{ts}.xlsx"

    wb.save(str(out_path))
    wb.close()

    print(f"  処理完了:")
    print(f"    ファイル名→URL対応: {len(fname_to_url)} 件")
    print(f"    G/H列置換: {replace_count} 箇所")
    print(f"    I列クリア: {clear_count} 行")
    print(f"    保存先: {out_path}")
    print()
    print(f"  ※ 元ファイル '{args.excel}' は変更されていません。")
    print("=" * 60)


if __name__ == "__main__":
    main()
